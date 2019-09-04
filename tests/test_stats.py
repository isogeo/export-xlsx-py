# -*- coding: UTF-8 -*-
#! python3

"""
    Usage from the repo root folder:

    ```python
    # for whole test
    python -m unittest tests.test_stats
    # for specific
    python -m unittest tests.test_stats.TestStats.test_translations_length
    ```
"""

# #############################################################################
# ########## Libraries #############
# ##################################

# Standard library
from datetime import date
import unittest

# 3rd party
from openpyxl import Workbook

# target
from isogeotoxlsx.utils import Stats


# #############################################################################
# ########## Classes ###############
# ##################################


class TestStats(unittest.TestCase):
    """Test stats and charts module."""

    def setUp(self):
        """Executed before each test."""
        pass

    def tearDown(self):
        """Executed after each test."""
        pass

    # -- TESTS ---------------------------------------------------
    def test_translations_length(self):
        """Ensure that different translations have the same length"""
        # this module
        app = Stats()
        # workbook
        wb = Workbook()

        # types of metadatas
        ws_types = wb.create_sheet(title="Types")
        app.md_types_repartition = {
            "raster": 50,
            "resource": 10,
            "service": 40,
            "vector": 100,
        }

        app.pie_types(ws_types)

        # formats of source datasets
        ws_formats = wb.create_sheet(title="Formats")
        app.li_data_formats = [
            "PostGIS",
            "WFS",
            "PostGIS",
            "WMS",
            "Esri Shapefiles",
            "Esri Shapefiles",
            "Esri Shapefiles",
            "Esri Shapefiles",
            "Esri Shapefiles",
        ]

        app.pie_formats(
            ws_formats,
            # cell_start_table="A"  # you can specify where to write table
        )

        # creation and modification dates
        ws_history = wb.create_sheet(title="History")
        app.li_dates_md_created = [
            date(2019, 1, 1),
            date(2019, 2, 1),
            date(2019, 1, 12),
            date(2019, 1, 12),
            date(2019, 1, 12),
            date(2019, 2, 14),
            date(2019, 2, 14),
            date(2019, 2, 14),
            date(2019, 2, 14),
            date(2019, 2, 28),
            date(2019, 3, 1),
            date(2019, 3, 2),
            date(2019, 3, 3),
            date(2019, 3, 4),
            date(2019, 3, 5),
            date(2019, 3, 5),
        ]

        app.li_dates_md_modified = [
            date(2019, 1, 1),
            date(2019, 2, 1),
            date(2019, 1, 12),
            date(2019, 3, 12),
            date(2019, 3, 12),
            date(2019, 3, 12),
            date(2019, 3, 12),
            date(2019, 3, 12),
            date(2019, 4, 12),
            date(2019, 2, 14),
            date(2019, 2, 28),
            date(2019, 3, 1),
            date(2019, 3, 2),
            date(2019, 3, 3),
            date(2019, 2, 4),
            date(2019, 2, 5),
        ]

        app.line_dates(ws=ws_history, cell_start_table="A1", cell_start_chart="E1")

        # write xlsx
        wb.save("test__unit_stats_charts.xlsx")


# #############################################################################
# ##### Stand alone program ########
# ##################################
if __name__ == "__main__":
    unittest.main()
