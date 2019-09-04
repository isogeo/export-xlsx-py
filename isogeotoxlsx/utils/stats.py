# -*- coding: UTF-8 -*-
#!/usr/bin/env python


# ----------------------------------------------------------------------------
# Name:         OpenCatalog to Excel
# Purpose:      Get metadatas from an Isogeo OpenCatlog and store it into
#               an Excel workbook.
#
# Author:       Isogeo
#
# Python:       2.7.x
# Created:      14/08/2014
# Updated:      28/01/2016
# ----------------------------------------------------------------------------

# ###########################################################################
# ########## Libraries ##########
# ###############################

# Standard library
from collections import Counter, defaultdict
import logging

# submodule
from isogeotoxlsx.i18n import I18N_EN, I18N_FR

# 3rd party library
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.worksheet import Worksheet


# ##############################################################################
# ############ Globals ############
# #################################

# LOG
logger = logging.getLogger("isogeotoxlsx")

# ############################################################################
# ######## Classes ###############
# ################################


class Stats(object):
    """Doc for Isogeo."""

    data_formats = []
    md_empty_fields = defaultdict(list)
    md_types_repartition = defaultdict(int)
    md_tags_occurences = defaultdict(int)

    def __init__(self, lang="fr"):
        """Instanciate stats class."""
        # self._ = _
        super(Stats, self).__init__()

        # LOCALE
        if lang.lower() == "fr":
            self.dates_fmt = "DD/MM/YYYY"
            self.locale_fmt = "fr_FR"
            self.tr = I18N_FR
        else:
            self.dates_fmt = "YYYY/MM/DD"
            self.locale_fmt = "uk_UK"
            self.tr = I18N_EN

    def attributes(self, ws_attributes: Worksheet, all_attributes: list):
        """Perform feature attributes analisis and write results into the
        wanted worksheet.

        :param Worksheet ws_attributes: sheet of a Workbook to write analisis
        :param list all_attributes: list of all feature attributes. It's a list of dicts.
        """
        idx_fa = 1
        # local arrays
        fa_names = []
        # fa_types = []
        # fa_alias = []
        # fa_descr = []

        # parsing
        for dico_fa in all_attributes:
            for fa in dico_fa:
                fa_names.append(fa.get("name"))
                # fa_alias.append(fa.get("alias", "NR"))
                # fa_types.append(fa.get("dataType"))
                # fa_descr.append(fa.get("description", "NR"))
                del fa

        # stats
        frq_names = Counter(fa_names)
        # frq_alias = Counter(fa_alias)
        # frq_types = Counter(fa_types)
        # frq_descr = Counter(fa_descr)

        # write
        ws = ws_attributes
        for fa in frq_names:
            idx_fa += 1
            ws["A{}".format(idx_fa)] = fa
            ws["B{}".format(idx_fa)] = frq_names.get(fa)

    def pie_formats(
        self,
        ws: Worksheet,
        li_formats: list = None,
        cell_start_table: str = "A20",
        cell_start_chart: str = "D20",
    ):
        """Calculates metadata types repartition and add a Pie chart to the wanted sheet of Workbook.

        :param Worksheet ws: sheet of a Workbook to write analisis
        :param list li_formats: list of all formats labels. If not specified, the class attribute will be used instaed
        :param str cell_start_table: cell of the sheet where to start writing table
        :param str cell_start_chart: cell of the sheet where to start writing the chart
        """
        if li_formats is None:
            li_formats = self.data_formats

        # build the data for pie chart
        data = Counter(li_formats)

        # get starting cells
        min_cell_start_table = ws[cell_start_table]

        # write headers
        ws.cell(
            row=min_cell_start_table.row,
            column=min_cell_start_table.column,
            value=self.tr.get("format"),
        )
        ws.cell(
            row=min_cell_start_table.row,
            column=min_cell_start_table.column + 1,
            value=self.tr.get("occurrences"),
        )

        # write data into worksheet
        row = min_cell_start_table.row
        for frmt, count in data.items():
            row += 1
            ws.cell(row=row, column=min_cell_start_table.column, value=frmt.title())
            ws.cell(row=row, column=min_cell_start_table.column + 1, value=count)

        # Pie chart
        pie = PieChart()
        labels = Reference(
            worksheet=ws,
            min_col=min_cell_start_table.column,
            min_row=min_cell_start_table.row + 1,
            max_row=row,
        )
        data = Reference(
            worksheet=ws,
            min_col=min_cell_start_table.column + 1,
            min_row=min_cell_start_table.row + 1,
            max_row=row,
        )
        pie.add_data(data)
        pie.set_categories(labels)
        pie.title = self.tr.get("format") + "s"

        # Cut the first slice out of the pie
        slice = DataPoint(idx=0, explosion=20)
        pie.series[0].data_points = [slice]

        ws.add_chart(pie, cell_start_chart)

    def pie_types(
        self,
        ws: Worksheet,
        types_counters: dict = None,
        cell_start_table: str = "A1",
        cell_start_chart: str = "D1",
    ):
        """Calculates metadata types repartition and add a Pie chart to the wanted sheet of Workbook.

        :param Worksheet ws: sheet of a Workbook to write analisis
        :param dict types_counters: dictionary of types/count. If not specified, the class attribute will be used instaed
        :param str cell_start_table: cell of the sheet where to start writing table
        :param str cell_start_chart: cell of the sheet where to start writing the chart
        """
        if types_counters is None:
            types_counters = self.md_types_repartition

        # build the data for pie chart
        data = (
            (self.tr.get("type"), self.tr.get("occurrences")),
            (self.tr.get("vector"), self.md_types_repartition.get("vector", 0)),
            (self.tr.get("raster"), self.md_types_repartition.get("raster", 0)),
            (self.tr.get("service"), self.md_types_repartition.get("service", 0)),
            (self.tr.get("resource"), self.md_types_repartition.get("resource", 0)),
        )

        # write data into worksheet
        for row in data:
            ws.append(row)

        # Pie chart
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=5)
        data = Reference(ws, min_col=2, min_row=1, max_row=5)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = self.tr.get("type") + "s"

        # Cut the first slice out of the pie
        slice = DataPoint(idx=0, explosion=20)
        pie.series[0].data_points = [slice]

        ws.add_chart(pie, cell_start_chart)


# ############################################################################
# ###### Stand alone program ########
# ###################################
if __name__ == "__main__":
    """Standalone execution and tests."""
    from openpyxl import Workbook

    # this module
    app = Stats()
    # workbook
    wb = Workbook()

    # types of metadatas
    ws_types = wb.create_sheet(title="Types")
    app.md_types_repartition = {
        "raster": 50,
        "resource": 10,
        "service": 40,
        "vector": 100,
    }

    app.pie_types(ws_types)

    # formats of source datasets
    ws_formats = wb.create_sheet(title="Formats")
    app.data_formats = [
        "PostGIS",
        "WFS",
        "PostGIS",
        "WMS",
        "Esri Shapefiles",
        "Esri Shapefiles",
        "Esri Shapefiles",
        "Esri Shapefiles",
        "Esri Shapefiles",
    ]

    app.pie_formats(
        ws_formats,
        # cell_start_table="A"  # you can specify where to write table
    )

    # write xlsx
    wb.save("test_stats_charts.xlsx")
