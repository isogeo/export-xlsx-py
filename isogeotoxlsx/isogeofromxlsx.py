# -*- coding: UTF-8 -*-
#! python3

"""
    Make Metadatas from Excel worksheet like those returned by Isogeo2xlsx.

"""

# ##############################################################################
# ########## Libraries #############
# ##################################

# Standard library
import logging
from pathlib import Path

# 3rd party library
from isogeo_pysdk import (
    IsogeoUtils,
    IsogeoChecker,
    Metadata,
    Keyword,
    Event,
    Contact
)
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

# custom submodules
from isogeotoxlsx.i18n import I18N_EN, I18N_FR
# from isogeotoxlsx.utils import Formatter, Stats

# ##############################################################################
# ############ Globals ############
# #################################

logger = logging.getLogger("isogeotoxlsx")
utils = IsogeoUtils()
checker = IsogeoChecker()

# root metadatas attributes to update
tup_easy_attr = (
    "abstract",
    "collectionContext",
    "collectionMethod",
    "created",
    "distance",
    "features",
    "format",
    "geometry",
    "language",
    "modified",
    "name",
    "path",
    "published",
    "scale",
    "title",
    "topologicalConsistency",
    "updateFrequency",
    "validFrom",
    "validTo",
    "validityComment"
)
# subressources attributes more complicated to update
tup_subressources_attr = (
    "conditions",
    "contacts",
    "coordinateSystem",
    "envelope",
    "events",
    "featureAttributes",
    "keywords",
    "limitations",
    "inspireThemes",
    "specifications",
)
# attributes filles or that we don't want to change value and field that we don't need 
tup_nogo_attr = (
    "_created",
    "_creator",
    "_id",
    "_modified",
    "featureAttributesCount",
    "hasLinkDownload",
    "hasLinkView",
    "hasLinkOther",
    "inspireConformance",
    "linkEdit",
    "linkView",
)

contacts_sheet_dict = {
    "zipCode": "contact_CP",
    "name": "contact",
    "email": "contact_mail",
    "phone": "contact_tel",
    "city": "contact_ville",
    "addressLine1": "contact_adresse"
}

# dict of inspire keyword ids for INSPIRe themes in french ## TO MOVE to i18n
dict_inspire_fr = {
    'Adresses': 'b181316d4e254c23839128062f914140',
    'Altitude': 'd73859f97ffb4d639c2ae4e8e60006b6',
    'Bâtiments': '57308eb7acb14320ae11bab71d12e9b5',
    'Caractéristiques géographiques météorologiques': 'f8ea6a2e5a9f44eabfb177186d6afd4b',
    'Caractéristiques géographiques océanographiques': '54f0b66d0dd14e7d8a1966c6accae69b',
    'Conditions atmosphériques': 'b7b896b9c46b4f569d53988e37413f60',
    'Dénominations géographiques': 'cb3414eb94114b4dbda1b1138ac65f66',
    'Géologie': '50db772e9689424ab110229da3d998a5',
    'Habitats et biotopes': '3c6cae4148ce47259456a7d1b1e196f5',
    'Hydrographie': '4bba97f69b0846609797e71690ca05b7',
    'Installations agricoles et aquacoles': '7bc9534cc5264f42af595f07b18dcb8f',
    'Installations de suivi environnemental': '54a04ed95bc540fb8779bd9790607fd7',
    'Lieux de production et sites industriels': '9ba4eeb2762242b79248b166c4a9a501',
    'Occupation des terres': '56bc50df59a8499eac2ea8edf82317c0',
    'Ortho-imagerie': '0399aa955b3a4f94973434ca83485b7c',
    'Parcelles cadastrales': 'e8a9612fc45b41baaf43862f4768fd44',
    'Ressources minérales': 'b27d27072ced49efa2a5b978698932ec',
    'Référentiels de coordonnées': 'ec246e0b891b4662843fe37053aaeec0',
    'Régions biogéographiques': '6e7061bd55d048e280b7ac7f2647b9e7',
    'Régions maritimes': 'c024243638e74eca990d53c51e49b900',
    'Répartition de la population — démographie': '6e81cdd6289e497291c5e404a277153a',
    'Répartition des espèces': '3f2f04b45d3d4c58ade588f31125afb4',
    'Réseaux de transport': '5aafe7f790e640e3bac7d168c9a4af21',
    'Santé et sécurité des personnes': '33ed965c15894a559adf43067ffd8c10',
    "Services d'utilité publique et services publics": '4456243f98e74a42a5b02944ca65ab94',
    'Sites protégés': '5a6c8cab0cdc429cb610907f9d13e7f2',
    'Sols': '0b34437444f64a728b5293ff4422dd68',
    "Sources d'énergie": 'e47ab302f0da4b1abc310cc7a55cb196',
    'Systèmes de maillage géographique': '25674aa9f458450e871bef8b0053603d',
    'Unités administratives': '5a64a5f464f94c55b9db1c99100fbd53',
    'Unités statistiques': '16c621ed81bc44d496b33b378c05c13a',
    'Usage des sols': 'c83ad1387c564061a722a244d874fb35',
    'Zones de gestion, de restriction ou de réglementation et unités de déclaration': 'fdff704c15aa4f90a3916395e8bbfd04',
    'Zones à risque naturel': '1dd88424fbad4b9a9eb12b71718833b8'
}
# dict of inspire keyword ids for INSPIRe themes in english ## TO MOVE to i18n
dict_inspire_en = {
    'Addresses': 'b181316d4e254c23839128062f914140',
    'Administrative units': '5a64a5f464f94c55b9db1c99100fbd53',
    'Agricultural and aquaculture facilities': '7bc9534cc5264f42af595f07b18dcb8f',
    'Area management/restriction/regulation zones and reporting units': 'fdff704c15aa4f90a3916395e8bbfd04',
    'Atmospheric conditions': 'b7b896b9c46b4f569d53988e37413f60',
    'Bio-geographical regions': '6e7061bd55d048e280b7ac7f2647b9e7',
    'Buildings': '57308eb7acb14320ae11bab71d12e9b5',
    'Cadastral parcels': 'e8a9612fc45b41baaf43862f4768fd44',
    'Coordinate reference systems': 'ec246e0b891b4662843fe37053aaeec0',
    'Elevation': 'd73859f97ffb4d639c2ae4e8e60006b6',
    'Energy resources': 'e47ab302f0da4b1abc310cc7a55cb196',
    'Environmental monitoring facilities': '54a04ed95bc540fb8779bd9790607fd7',
    'Geographical grid systems': '25674aa9f458450e871bef8b0053603d',
    'Geographical names': 'cb3414eb94114b4dbda1b1138ac65f66',
    'Geology': '50db772e9689424ab110229da3d998a5',
    'Habitats and biotopes': '3c6cae4148ce47259456a7d1b1e196f5',
    'Human health and safety': '33ed965c15894a559adf43067ffd8c10',
    'Hydrography': '4bba97f69b0846609797e71690ca05b7',
    'Land cover': '56bc50df59a8499eac2ea8edf82317c0',
    'Land use': 'c83ad1387c564061a722a244d874fb35',
    'Meteorological geographical features': 'f8ea6a2e5a9f44eabfb177186d6afd4b',
    'Mineral resources': 'b27d27072ced49efa2a5b978698932ec',
    'Natural risk zones': '1dd88424fbad4b9a9eb12b71718833b8',
    'Oceanographic geographical features': '54f0b66d0dd14e7d8a1966c6accae69b',
    'Orthoimagery': '0399aa955b3a4f94973434ca83485b7c',
    'Population distribution — demography': '6e81cdd6289e497291c5e404a277153a',
    'Production and industrial facilities': '9ba4eeb2762242b79248b166c4a9a501',
    'Protected sites': '5a6c8cab0cdc429cb610907f9d13e7f2',
    'Sea regions': 'c024243638e74eca990d53c51e49b900',
    'Soil': '0b34437444f64a728b5293ff4422dd68',
    'Species distribution': '3f2f04b45d3d4c58ade588f31125afb4',
    'Statistical units': '16c621ed81bc44d496b33b378c05c13a',
    'Transport networks': '5aafe7f790e640e3bac7d168c9a4af21',
    'Utility and governmental services': '4456243f98e74a42a5b02944ca65ab94'
}

# ##############################################################################
# ########## Classes ###############
# ##################################


class IsogeoFromxlsx:
    """Used to read Isogeo Metadata stored into an Excel worksheet (.xlsx)

    :param Path file_path: the path of xlsx file to read
    :param str lang: selected language for input
    """

    def __init__(
        self, file_path: Path = "", lang: str = "fr",
    ):
        """Instanciating the output workbook.

        :param str lang: selected language for output
        :param str url_base_edit: base url to format edit links (basically app.isogeo.com)
        :param str url_base_view: base url to format view links (basically open.isogeo.com)
        """
        super(IsogeoFromxlsx, self).__init__()

        self.file_path = Path(file_path)
        # Initiate reading
        if self.file_path.is_file():
            logger.info("Read the excel file {}".format(file_path))
            self.work_book = load_workbook(
                filename=self.file_path, read_only=True
            )
        else:
            raise ValueError(
                "'{}' does not exist or is not reachable.".format(file_path)
            )

        self.li_ignored_md = []
        self.md_read = []
        self.contacts_read = []

        s_date = NamedStyle(name="date")
        self.worksheets_dict = {}
        # LOCALE
        self.lang = lang
        if self.lang.lower() == "fr":
            s_date.number_format = "dd/mm/yyyy"
            self.dates_fmt = "DD/MM/YYYY"
            self.locale_fmt = "fr_FR"
            self.tr = I18N_FR
            self.dict_inspire = dict_inspire_fr
            self.worksheets_dict = {
                self.tr.get("vector"): None,
                "Raster": None,
                "Services": None,
                "Contacts": None,
            }

        else:
            s_date.number_format = "yyyy/mm/dd"
            self.dates_fmt = "YYYY/MM/DD"
            self.locale_fmt = "uk_UK"
            self.tr = I18N_EN
            self.dict_inspire = dict_inspire_en
            self.worksheets_dict = {
                self.tr.get("vector"): None,
                "Raster": None,
                "Services": None,
                "Contacts": None,
            }

        for sheet in self.worksheets_dict:
            if sheet in self.work_book.sheetnames:
                self.worksheets_dict[sheet] = self.work_book[sheet]
            else:
                logger.info("No '{}' sheet found in the file".format(sheet))

        if self.worksheets_dict.get("Contacts"):
            self.retrieve_contacts()
        else:
            logger.debug("No contacts to retrieve")

    def read_file(self):
        self.retrieve_vector_metadatas()
        logger.info("{} contacts and {} metadatas read".format(len(self.contacts_read), len(self.md_read)))
        if len(self.li_ignored_md) > 0:
            logger.info("{} metadatas ignored because of invalid UUID".format(len(self.li_ignored_md)))
        else:
            pass

    def retrieve_vector_metadatas(self):
        logger.debug("Retrieving vector metadatas from '{}' sheet".format(self.tr.get("vector")))
        ws_vectors = self.worksheets_dict.get(self.tr.get("vector"))
        attr_index_dict = self.build_index_dict(work_sheet=ws_vectors, ref_dict=self.tr)
        # browsing the frame starting at the second line because we don't nead to read headers
        for row in ws_vectors.iter_rows(min_row=2):
            # retrieve metadata id and check UUID validity
            md_uuid = row[attr_index_dict.get("_id")].value
            if checker.check_is_uuid(md_uuid):
                md_dict = {
                    "md": None,
                    "conditions": [],
                    "contacts": [],
                    "coordinateSystem": {},
                    "envelope": {},
                    "events": [],
                    "featureAttributes": [],
                    "keywords": [],
                    "limitations": [],
                    "specifications": [],
                    "inspireThemes": {},
                }
                # create Metadata object
                md = Metadata()
                md._id = md_uuid
                # for attr in Metadata().ATTR_TYPES:
                #     if attr not in list(attr_index_dict.keys()):
                #         if Metadata().ATTR_TYPES[attr] == str:
                #             setattr(md, attr, "***nogo***")
                #         elif Metadata().ATTR_TYPES[attr] == list:
                #             pass
                #         elif Metadata().ATTR_TYPES[attr] == dict:
                #             pass
                #         else:
                #             setattr(md, attr, -1)
                #     else:
                #         pass
                for field in attr_index_dict:
                    # retieve the value taken by the row in this field
                    field_value = row[attr_index_dict.get(field)].value
                    if field_value:
                        # specifics fields and fields corresponding to attributes that we don't want to change value
                        if field in tup_nogo_attr:
                            pass
                            # if hasattr(md, field):
                            #     try:
                            #         setattr(md, field, "***nogo***")
                            #     except Exception as e:
                            #         logger.debug("'{}' attribute can't be set :{}".format(field, e))
                            # else:
                            #     logger.debug("'{}' field is not needed".format(field))
                            #     pass
                        # for root Metadata attributes that are not filled by the scan and that exist in the frame
                        elif field in tup_easy_attr:
                            try:
                                if Metadata().ATTR_TYPES[field] == str:
                                    if field != "validFrom" and field != "validTo":
                                        setattr(md, field, str(field_value))
                                    else:
                                        setattr(md, field, field_value.strftime("%Y-%m-%dT%H:%M:%S+00:00"))
                                elif Metadata().ATTR_TYPES[field] == int:
                                    setattr(md, field, int(field_value))
                                else:
                                    setattr(md, field, float(field_value))
                            except Exception as e:
                                logger.debug(
                                    "'{}' attribute can't be set: {}".format(field, e)
                                )
                        # sub ressources
                        elif field in tup_subressources_attr:
                            if field == "keywords":
                                md_dict["keywords"] = self.build_keywords(field_value)
                            elif field == "inspireThemes":
                                md_dict["inspireThemes"] = self.build_inspireTh(field_value)
                            elif field == "created":
                                md_dict["events"].append(self.build_event(event_date=field_value, kind="creation"))
                            elif field == "published":
                                md_dict["events"].append(self.build_event(event_date=field_value, kind="publication"))
                            elif field == "modified":
                                md_dict["events"].append(self.build_event(event_date=field_value, kind="update"))
                            elif field == "contacts":
                                if len(self.contacts_read) > 0:
                                    md_dict["contacts"] = self.build_md_contacts(field_value)
                                else:
                                    logger.debug(
                                        "'{}' metadata has contact(s) set in '{}' sheet but no contact is referenced in the 'Contacts' sheet".format(
                                            md._id,
                                            self.tr.get("vector")
                                        )
                                    )
                            elif field == "conditions":
                                md_dict["conditions"] = field_value
                            else:
                                pass
                        else:
                            logger.warning(
                                "'{}' field is not expected".format(
                                    field
                                )
                            )
                            continue
                    else:
                        pass
                md_dict["md"] = md
                self.md_read.append(md_dict)
            else:
                logger.info("'{}' is not a valid UUID")
                self.li_ignored_md.append(md_uuid)
                continue
        return

    def build_index_dict(self, work_sheet, ref_dict: dict):
        attr_index_dict = {}
        for i in range(1, work_sheet.max_column + 1):
            # retrieve column header
            header = work_sheet.cell(row=1, column=i).value
            # retrieve metadata attribute corresponding to the header
            if header in list(ref_dict.values()):
                attribute = [k for k, v in ref_dict.items() if v == header][0]
                # put metadata attribute as key and corresponding column index as value
                attr_index_dict[attribute] = i - 1
            else:
                logger.warning("'{}' is not a regular column name".format(header))
        return attr_index_dict

    def build_list(self, text: str):
        li_values = []
        for value in text.split(";"):
            li_values.append(value.strip())

        return li_values

    def build_keywords(self, keywords_value: str):
        li_kw = self.build_list(keywords_value)
        li_isogeo_kw = []
        for kw in li_kw:
            dict_kw = Keyword().to_dict_creation()
            dict_kw["text"] = kw
            isogeo_kw = Keyword(**dict_kw)
            li_isogeo_kw.append(isogeo_kw)

        return li_isogeo_kw

    def build_inspireTh(self, inspireTh_value: str):
        li_th = self.build_list(inspireTh_value)
        dict_isogeo_th = {}
        for th in li_th:
            if th in list(self.dict_inspire.keys()):
                dict_isogeo_th[th] = self.dict_inspire.get(th)
            else:
                logger.debug("Unexpected INSPIRE theme found in the file : '{}'".format(th))

        return dict_isogeo_th

    def build_event(self, event_date: str, kind: str):
        dict_event = Event().to_dict_creation()

        dict_event["date"] = event_date.strftime("%Y-%m-%dT%H:%M:%S+00:00")
        dict_event["kind"] = kind

        creation_event = Event(**dict_event)

        return creation_event

    def build_md_contacts(self, contacts_value: str):
        li_ctcts = self.build_list(contacts_value)
        li_emails_ctct = []
        for ctct in li_ctcts:
            isogeo_contact = [contact for contact in self.contacts_read if contact.email in ctct]
            if len(isogeo_contact) == 0:
                logger.debug("'{}' contact is not referenced in 'Contacts' sheet".format(ctct))
            else:
                li_emails_ctct.append(isogeo_contact[0])

        return li_emails_ctct

    def retrieve_contacts(self):
        logger.debug("Retrieving contacts from 'Contacts' sheet")
        ws_contacts = self.worksheets_dict.get("Contacts")
        attr_index_dict = self.build_index_dict(work_sheet=ws_contacts, ref_dict=contacts_sheet_dict)
        # browsing the frame starting at the second line because we don't nead to read headers
        for row in ws_contacts.iter_rows(min_row=2):
            contact_dict = {
                "zipCode": "",
                "name": "",
                "email": "",
                "phone": "",
                "city": "",
                "addressLine1": ""
            }
            for field in attr_index_dict:
                field_value = row[attr_index_dict.get(field)].value
                contact_dict[field] = field_value
            self.contacts_read.append(Contact(**contact_dict))
        return